#!/usr/bin/env python3
"""
gen_testcase_excel.py — 将测试用例数据生成 Excel 文件

用法：
    python gen_testcase_excel.py <input_json> <output.xlsx>

input_json 结构：
{
  "title": "功能名称",
  "modules": [
    {
      "name": "1 入口&首页",
      "cases": [
        {
          "id": "TC-101",
          "title": "点击入口成功跳转",
          "priority": "P0",
          "precondition": "已登录，网络正常",
          "steps": ["1. 打开APP", "2. 点击xxx入口"],
          "expected": "成功跳转到目标页，页面正常展示",
          "module": "1 入口&首页"
        }
      ]
    }
  ]
}

每个功能模块生成一个 Sheet，Sheet 名为模块名（最多31字符）。
依赖：openpyxl
"""

import json
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("缺少依赖，正在安装 openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


HEADERS = ["ID", "标题", "优先级", "所属模块", "前置条件", "操作步骤", "预期结果"]
COL_WIDTHS = [12, 28, 8, 20, 25, 40, 30]

PRIORITY_COLORS = {
    "P0": "FFE0E0",  # 淡红
    "P1": "FFF3CD",  # 淡黄
    "P2": "E8F5E9",  # 淡绿
}

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)


def thin_border():
    side = Side(style="thin", color="BBBBBB")
    return Border(left=side, right=side, top=side, bottom=side)


def write_sheet(ws, module_name, cases):
    # 写表头
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()

    ws.row_dimensions[1].height = 22

    # 写数据行
    for row_idx, case in enumerate(cases, 2):
        priority = case.get("priority", "P2")
        fill_color = PRIORITY_COLORS.get(priority, "FFFFFF")
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        steps = case.get("steps", [])
        if isinstance(steps, list):
            steps_text = "\n".join(steps)
        else:
            steps_text = str(steps)

        values = [
            case.get("id", ""),
            case.get("title", ""),
            priority,
            case.get("module", module_name),
            case.get("precondition", ""),
            steps_text,
            case.get("expected", ""),
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border()
            cell.font = Font(size=10)

        ws.row_dimensions[row_idx].height = max(30, len(steps_text.splitlines()) * 15)

    # 设置列宽
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 冻结首行
    ws.freeze_panes = "A2"


def create_excel_file(data, output_path):
    title = data.get("title", "测试用例")
    modules = data.get("modules", [])

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 移除默认空 Sheet

    total_cases = 0

    for mod in modules:
        module_name = mod.get("name", "未命名模块")
        cases = mod.get("cases", [])
        if not cases:
            continue

        # Sheet 名称最多31字符，去掉非法字符
        sheet_name = module_name[:31].replace("/", "／").replace("\\", "＼").replace("?", "？").replace("*", "＊").replace("[", "【").replace("]", "】").replace(":", "：")
        ws = wb.create_sheet(title=sheet_name)
        write_sheet(ws, module_name, cases)
        total_cases += len(cases)

    if not wb.sheetnames:
        # 没有任何数据，创建一个空 Sheet
        wb.create_sheet(title=title[:31])

    wb.save(output_path)
    print(f"✅ Excel 文件已生成：{output_path}（共 {total_cases} 条用例，{len(wb.sheetnames)} 个 Sheet）")
    return output_path


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("用法：python gen_testcase_excel.py <input.json> <output.xlsx>")
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2]

    with open(input_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    create_excel_file(data, output_path)
